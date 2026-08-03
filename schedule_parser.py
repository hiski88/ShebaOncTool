from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable, Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.utils import get_column_letter

from core import HEBREW_MONTHS, hebrew_weekday, important_day_name, normalize_spaces

MONTH_NAME_TO_NUMBER = {name: number for number, name in HEBREW_MONTHS.items()}
MONTH_NAME_TO_NUMBER.update({"מרס": 3})


@dataclass
class CellData:
    row: int
    col: int
    coordinate: str
    value: Any
    text: str
    strike_ranges: list[tuple[int, int]] = field(default_factory=list)

    def is_struck(self, start: int, end: int) -> bool:
        if end <= start:
            return False
        overlap = 0
        for range_start, range_end in self.strike_ranges:
            overlap += max(0, min(end, range_end) - max(start, range_start))
        return overlap >= max(1, (end - start) / 2)


@dataclass
class SheetData:
    name: str
    nrows: int
    ncols: int
    cells: dict[tuple[int, int], CellData]

    def cell(self, row: int, col: int) -> CellData:
        return self.cells.get((row, col), CellData(row, col, f"{get_column_letter(col + 1)}{row + 1}", None, "", []))


@dataclass
class WorkbookData:
    sheets: list[SheetData]

    def sheet_by_preference(self, names: Sequence[str]) -> SheetData:
        for requested in names:
            for sheet in self.sheets:
                if normalize_spaces(sheet.name) == normalize_spaces(requested):
                    return sheet
        return self.sheets[0]


@dataclass(frozen=True)
class PersonOccurrence:
    name: str
    start: int
    end: int
    struck: bool
    segment_after: str


def _value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "כן" if value else "לא"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _xlsx_cell_data(cell) -> CellData:
    value = cell.value
    strike_ranges: list[tuple[int, int]] = []
    if isinstance(value, CellRichText):
        parts: list[str] = []
        cursor = 0
        for part in value:
            if isinstance(part, TextBlock):
                text = str(part.text)
                struck = bool(getattr(part.font, "strike", False))
            else:
                text = str(part)
                struck = bool(getattr(cell.font, "strike", False))
            parts.append(text)
            if struck and text:
                strike_ranges.append((cursor, cursor + len(text)))
            cursor += len(text)
        text_value = "".join(parts)
        raw_value = text_value
    else:
        raw_value = value
        text_value = _value_to_text(value)
        if text_value and bool(getattr(cell.font, "strike", False)):
            strike_ranges.append((0, len(text_value)))
    return CellData(cell.row - 1, cell.column - 1, cell.coordinate, raw_value, text_value, strike_ranges)


def _read_xlsx(file_bytes: bytes) -> WorkbookData:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, rich_text=True)
    sheets: list[SheetData] = []
    for worksheet in workbook.worksheets:
        cells: dict[tuple[int, int], CellData] = {}
        for row in worksheet.iter_rows():
            for cell in row:
                data = _xlsx_cell_data(cell)
                if data.value is not None or data.text or data.strike_ranges:
                    cells[(data.row, data.col)] = data
        sheets.append(SheetData(worksheet.title, worksheet.max_row, worksheet.max_column, cells))
    return WorkbookData(sheets)


def _read_xls(file_bytes: bytes) -> WorkbookData:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise RuntimeError("קריאת קובצי XLS דורשת את החבילה xlrd") from exc

    workbook = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
    sheets: list[SheetData] = []
    for worksheet in workbook.sheets():
        cells: dict[tuple[int, int], CellData] = {}
        for row in range(worksheet.nrows):
            for col in range(worksheet.ncols):
                cell = worksheet.cell(row, col)
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(cell.value, workbook.datemode)
                text = _value_to_text(value)
                strike_ranges: list[tuple[int, int]] = []
                if text:
                    default_font_index = 0
                    try:
                        default_font_index = workbook.xf_list[cell.xf_index].font_index
                    except Exception:
                        pass
                    run_list = list(worksheet.rich_text_runlist_map.get((row, col), []))
                    runs: list[tuple[int, int]] = []
                    if not run_list or run_list[0][0] != 0:
                        runs.append((0, default_font_index))
                    runs.extend((int(offset), int(font_index)) for offset, font_index in run_list)
                    runs = sorted(dict(runs).items())
                    for index, (offset, font_index) in enumerate(runs):
                        end = runs[index + 1][0] if index + 1 < len(runs) else len(text)
                        try:
                            struck = bool(workbook.font_list[font_index].struck_out)
                        except Exception:
                            struck = False
                        if struck and end > offset:
                            strike_ranges.append((offset, end))
                if value not in (None, "") or strike_ranges:
                    coordinate = f"{get_column_letter(col + 1)}{row + 1}"
                    cells[(row, col)] = CellData(row, col, coordinate, value, text, strike_ranges)
        sheets.append(SheetData(worksheet.name, worksheet.nrows, worksheet.ncols, cells))
    return WorkbookData(sheets)


def read_schedule_workbook(file_bytes: bytes, filename: str) -> WorkbookData:
    lower = filename.lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        return _read_xls(file_bytes)
    if lower.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(file_bytes)
    raise ValueError("יש להעלות קובץ XLS או XLSX")


def _parse_date_value(value: Any, month_year: tuple[int, int] | None) -> date | None:
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        integer = int(value)
        if 1 <= integer <= 31 and month_year:
            year, month = month_year
            try:
                return date(year, month, integer)
            except ValueError:
                return None
    text = normalize_spaces(value)
    if not text:
        return None
    for day_first in (True, False):
        try:
            parsed = pd.to_datetime(text, dayfirst=day_first, errors="raise")
            if 2000 <= parsed.year <= 2100:
                return parsed.date()
        except Exception:
            pass
    if text.isdigit() and month_year:
        integer = int(text)
        if 1 <= integer <= 31:
            year, month = month_year
            try:
                return date(year, month, integer)
            except ValueError:
                return None
    return None


def detect_month_year(sheet: SheetData, config: Mapping[str, Any]) -> tuple[int, int] | None:
    title_cells = config.get("schedule", {}).get("title_cells", [])
    coordinates = {cell.coordinate: cell for cell in sheet.cells.values()}
    texts: list[str] = []
    for coordinate in title_cells:
        if coordinate in coordinates:
            texts.append(coordinates[coordinate].text)
    for row in range(min(sheet.nrows, 6)):
        for col in range(min(sheet.ncols, 10)):
            texts.append(sheet.cell(row, col).text)

    month_pattern = "|".join(sorted((re.escape(name) for name in MONTH_NAME_TO_NUMBER), key=len, reverse=True))
    pattern = re.compile(rf"({month_pattern})\s*[-/]?\s*(20\d{{2}})")
    for text in texts:
        match = pattern.search(normalize_spaces(text))
        if match:
            return int(match.group(2)), MONTH_NAME_TO_NUMBER[match.group(1)]

    for row in range(min(sheet.nrows, 50)):
        parsed = _parse_date_value(sheet.cell(row, 0).value, None)
        if parsed:
            return parsed.year, parsed.month
    return None


def _has_modern_holiday_column(sheet: SheetData) -> bool:
    for row in range(min(sheet.nrows, 12)):
        text = normalize_spaces(sheet.cell(row, 2).text)
        if "חג" in text or "יום מיוחד" in text:
            return True
    return False


def _effective_column(index: int, modern_layout: bool) -> int:
    return index + 1 if modern_layout and index >= 2 else index


def schedule_rows(sheet: SheetData, config: Mapping[str, Any]) -> list[tuple[int, date]]:
    month_year = detect_month_year(sheet, config)
    rows: list[tuple[int, date]] = []
    for row in range(sheet.nrows):
        current = _parse_date_value(sheet.cell(row, 0).value, month_year)
        if current:
            rows.append((row, current))
    return rows


def _clean_candidate(fragment: str) -> str:
    fragment = fragment.strip(" \t\n,;:|()[]{}")
    return normalize_spaces(fragment)


def infer_employee_names(workbook: WorkbookData, config: Mapping[str, Any]) -> list[str]:
    sheet = workbook.sheet_by_preference(config.get("schedule", {}).get("sheet_names", []))
    modern = _has_modern_holiday_column(sheet)
    excluded = {normalize_spaces(term).casefold() for term in config.get("non_name_terms", [])}
    columns = config.get("schedule", {}).get("columns", [])
    counts: dict[str, int] = {}
    duty_names: set[str] = set()

    for column in columns:
        col = _effective_column(int(column["index"]), modern)
        if col >= sheet.ncols:
            continue
        for row, _ in schedule_rows(sheet, config):
            text = sheet.cell(row, col).text
            if not text:
                continue
            fragments = re.split(r"[\n,;]+", text)
            for fragment in fragments:
                candidate = _clean_candidate(fragment)
                folded = candidate.casefold()
                if not candidate or folded in excluded:
                    continue
                if any(term and term in folded for term in excluded):
                    continue
                if any(char.isdigit() for char in candidate):
                    continue
                if "/" in candidate or "+" in candidate:
                    continue
                if len(candidate) < 2 or len(candidate) > 35:
                    continue
                if len(candidate.split()) > 4:
                    continue
                if not re.fullmatch(r"[א-תA-Za-zÀ-ÖØ-öø-ÿ'׳״\- ]+", candidate):
                    continue
                counts[candidate] = counts.get(candidate, 0) + 1
                if column.get("kind") == "duty":
                    duty_names.add(candidate)

    names = [name for name, count in counts.items() if count >= 2 or name in duty_names]
    return sorted(names, key=lambda item: (item.casefold(), item))


def _find_occurrences(cell: CellData, employee_names: Sequence[str]) -> list[PersonOccurrence]:
    matches: list[tuple[int, int, str]] = []
    occupied: list[tuple[int, int]] = []
    for name in sorted({normalize_spaces(item) for item in employee_names if normalize_spaces(item)}, key=len, reverse=True):
        pattern_body = r"\s+".join(re.escape(part) for part in name.split())
        pattern = re.compile(rf"(?<![\w]){pattern_body}(?![\w])", re.UNICODE)
        for match in pattern.finditer(cell.text):
            if any(max(match.start(), start) < min(match.end(), end) for start, end in occupied):
                continue
            matches.append((match.start(), match.end(), name))
            occupied.append((match.start(), match.end()))
    matches.sort()

    occurrences: list[PersonOccurrence] = []
    for index, (start, end, name) in enumerate(matches):
        next_start = matches[index + 1][0] if index + 1 < len(matches) else len(cell.text)
        segment = cell.text[end:next_start].strip(" \t\n,;:|-")
        occurrences.append(PersonOccurrence(name, start, end, cell.is_struck(start, end), segment))
    return occurrences


def _aliases_in_text(text: str, config: Mapping[str, Any]) -> list[dict[str, Any]]:
    folded = normalize_spaces(text).casefold()
    matches: list[tuple[int, int, dict[str, Any]]] = []
    for alias in config.get("activity_aliases", []):
        for term in alias.get("terms", []):
            normalized_term = normalize_spaces(term).casefold()
            if normalized_term and normalized_term in folded:
                matches.append((len(normalized_term), folded.find(normalized_term), alias))
    matches.sort(key=lambda item: (-item[0], item[1]))
    unique: list[dict[str, Any]] = []
    seen: set[str] = set()
    for _, _, alias in matches:
        code = str(alias.get("code"))
        if code not in seen:
            seen.add(code)
            unique.append(alias)
    return unique


def _clinic_subtype(segment: str, config: Mapping[str, Any]) -> str:
    aliases = []
    for alias in config.get("activity_aliases", []):
        for term in alias.get("terms", []):
            aliases.append(normalize_spaces(term))
    subtype = segment
    for term in sorted((item for item in aliases if item), key=len, reverse=True):
        subtype = re.sub(re.escape(term), "", subtype, flags=re.IGNORECASE)
    return normalize_spaces(subtype.strip(" /,+-|"))


def _base_record(
    current: date,
    employee: str,
    cell: CellData,
    source: Mapping[str, Any],
    holiday: str,
) -> dict[str, Any]:
    return {
        "date": current,
        "day": hebrew_weekday(current),
        "holiday": holiday,
        "employee": employee,
        "source_code": str(source.get("source_code", "")),
        "source_label": str(source.get("label", "")),
        "source_kind": str(source.get("kind", "")),
        "slot": int(source.get("slot", 1)),
        "source_cell": cell.coordinate,
        "raw_text": cell.text,
    }


def parse_schedule(
    workbook: WorkbookData,
    config: Mapping[str, Any],
    employee_names: Sequence[str],
) -> pd.DataFrame:
    sheet = workbook.sheet_by_preference(config.get("schedule", {}).get("sheet_names", []))
    modern = _has_modern_holiday_column(sheet)
    rows = schedule_rows(sheet, config)
    records: list[dict[str, Any]] = []
    task_labels = config.get("task_labels", {})

    for row_index, current in rows:
        holiday = normalize_spaces(sheet.cell(row_index, 2).text) if modern else important_day_name(current)
        for source in config.get("schedule", {}).get("columns", []):
            col = _effective_column(int(source["index"]), modern)
            if col >= sheet.ncols:
                continue
            cell = sheet.cell(row_index, col)
            if not cell.text:
                continue
            occurrences = _find_occurrences(cell, employee_names)
            if not occurrences:
                continue

            for occurrence in occurrences:
                base = _base_record(current, occurrence.name, cell, source, holiday)
                aliases = _aliases_in_text(occurrence.segment_after, config)
                status_aliases = [alias for alias in aliases if alias.get("kind") == "status"]
                task_aliases = [alias for alias in aliases if alias.get("kind") == "task"]

                if source.get("kind") == "status":
                    if status_aliases:
                        for alias in status_aliases:
                            code = str(alias["code"])
                            records.append(
                                {
                                    **base,
                                    "record_type": "status",
                                    "task_code": code,
                                    "task_label": str(alias.get("label", task_labels.get(code, code))),
                                    "subtype": "",
                                    "struck": occurrence.struck,
                                }
                            )
                    else:
                        records.append(
                            {
                                **base,
                                "record_type": "status",
                                "task_code": "absence_other",
                                "task_label": task_labels.get("absence_other", "היעדרויות / אחר"),
                                "subtype": normalize_spaces(occurrence.segment_after),
                                "struck": occurrence.struck,
                            }
                        )
                    continue

                if occurrence.struck:
                    if status_aliases:
                        for alias in status_aliases:
                            code = str(alias["code"])
                            records.append(
                                {
                                    **base,
                                    "record_type": "status",
                                    "task_code": code,
                                    "task_label": str(alias.get("label", task_labels.get(code, code))),
                                    "subtype": "",
                                    "struck": True,
                                }
                            )
                    elif task_aliases:
                        labels = ", ".join(str(alias.get("label", alias.get("code"))) for alias in task_aliases)
                        records.append(
                            {
                                **base,
                                "record_type": "status",
                                "task_code": "moved_assignment",
                                "task_label": "שובץ/ה בתחנה אחרת",
                                "subtype": labels,
                                "struck": True,
                            }
                        )
                    else:
                        records.append(
                            {
                                **base,
                                "record_type": "status",
                                "task_code": "struck_out",
                                "task_label": "שם מחוק",
                                "subtype": "",
                                "struck": True,
                            }
                        )
                    continue

                if status_aliases:
                    for alias in status_aliases:
                        code = str(alias["code"])
                        records.append(
                            {
                                **base,
                                "record_type": "status",
                                "task_code": code,
                                "task_label": str(alias.get("label", task_labels.get(code, code))),
                                "subtype": "",
                                "struck": False,
                            }
                        )
                    continue

                if task_aliases:
                    alias = task_aliases[0]
                    code = str(alias["code"])
                    if code == "department_regular":
                        code = "department"
                    records.append(
                        {
                            **base,
                            "record_type": "task",
                            "task_code": code,
                            "task_label": str(alias.get("label", task_labels.get(code, code))),
                            "subtype": _clinic_subtype(occurrence.segment_after, config) if code == "clinic" else "",
                            "struck": False,
                        }
                    )
                    continue

                if source.get("kind") == "tracking":
                    continue

                code = str(source.get("source_code"))
                subtype = ""
                if code == "clinic":
                    subtype = normalize_spaces(occurrence.segment_after)
                records.append(
                    {
                        **base,
                        "record_type": "task",
                        "task_code": code,
                        "task_label": str(task_labels.get(code, source.get("label", code))),
                        "subtype": subtype,
                        "struck": False,
                    }
                )

    if not records:
        empty = pd.DataFrame(
            columns=[
                "date", "day", "holiday", "employee", "record_type", "task_code", "task_label", "subtype",
                "source_code", "source_label", "source_kind", "slot", "source_cell", "raw_text", "struck",
            ]
        )
        empty.attrs["schedule_dates"] = [current for _, current in rows]
        return empty

    dataframe = pd.DataFrame(records)
    dataframe = _classify_weekend_department_and_duties(dataframe, task_labels)
    priority = {"station": 0, "duty": 0, "on_call": 0, "status": 0, "tracking": 1}
    dataframe["_priority"] = dataframe["source_kind"].map(priority).fillna(0)
    dataframe = dataframe.sort_values(["date", "employee", "record_type", "task_code", "_priority", "source_cell"])
    task_mask = dataframe["record_type"] == "task"
    tasks = dataframe[task_mask].drop_duplicates(["date", "employee", "task_code", "subtype"], keep="first")
    statuses = dataframe[~task_mask].drop_duplicates(["date", "employee", "task_code", "subtype"], keep="first")
    dataframe = pd.concat([tasks, statuses], ignore_index=True).drop(columns=["_priority"], errors="ignore")
    dataframe = dataframe.sort_values(["date", "employee", "record_type", "task_code"]).reset_index(drop=True)
    dataframe.attrs["schedule_dates"] = [current for _, current in rows]
    return dataframe


def _classify_weekend_department_and_duties(dataframe: pd.DataFrame, task_labels: Mapping[str, str]) -> pd.DataFrame:
    result = dataframe.copy()
    task_rows = result["record_type"] == "task"

    ward_duty_mask = task_rows & (result["task_code"] == "ward_duty")
    for index, row in result[ward_duty_mask].iterrows():
        weekday = row["date"].weekday()
        if weekday == 4:
            code = "ward_duty_friday"
        elif weekday == 5:
            code = "ward_duty_saturday"
        else:
            code = "ward_duty_regular"
        result.at[index, "task_code"] = code
        result.at[index, "task_label"] = task_labels.get(code, code)

    department_mask = task_rows & (result["task_code"] == "department")
    for current in sorted(result.loc[department_mask, "date"].unique()):
        current_date = current.date() if isinstance(current, pd.Timestamp) else current
        day_indices = result.index[department_mask & (result["date"] == current)].tolist()
        current_duty_code = "ward_duty_friday" if current_date.weekday() == 4 else "ward_duty_saturday" if current_date.weekday() == 5 else "ward_duty_regular"
        current_duty_names = set(
            result.loc[
                (result["date"] == current) & (result["record_type"] == "task") & (result["task_code"] == current_duty_code),
                "employee",
            ]
        )
        previous_friday_names: set[str] = set()
        if current_date.weekday() == 5:
            previous = current_date.fromordinal(current_date.toordinal() - 1)
            previous_friday_names = set(
                result.loc[
                    (result["date"] == previous)
                    & (result["record_type"] == "task")
                    & (result["task_code"] == "ward_duty_friday"),
                    "employee",
                ]
            )

        for index in day_indices:
            employee = result.at[index, "employee"]
            if current_date.weekday() == 4:
                code = "department_weekend_friday_duty" if employee in current_duty_names else "friday_visit"
            elif current_date.weekday() == 5:
                if employee in current_duty_names:
                    code = "department_weekend_saturday_duty"
                elif employee in previous_friday_names:
                    code = "department_weekend_outgoing"
                else:
                    code = "department_weekend"
            else:
                code = "department_regular"
            result.at[index, "task_code"] = code
            result.at[index, "task_label"] = task_labels.get(code, code)
    return result


SUMMARY_DEFINITIONS: list[tuple[str, set[str]]] = [
    ("מחלקה יום רגיל", {"department_regular"}),
    (
        "מחלקה סופ\"ש",
        {
            "department_weekend", "department_weekend_friday_duty", "friday_visit",
            "department_weekend_outgoing", "department_weekend_saturday_duty",
        },
    ),
    ("ביקור שישי", {"friday_visit"}),
    ("תורנות מחלקה רגילה", {"ward_duty_regular"}),
    ("תורנות שישי", {"ward_duty_friday"}),
    ("תורנות שבת", {"ward_duty_saturday"}),
    ("תורנות מיון", {"er_duty"}),
    ("תורנות אשפוז יום", {"day_hospital_duty"}),
    ("מיון יום", {"oncology_er"}),
    ("אשפוז יום", {"day_hospital"}),
    ("אבחון מהיר", {"rapid_diagnosis"}),
    ("תגבור מיון", {"er_reinforcement"}),
    ("מרפאות", {"clinic"}),
    ("כוננויות", {"ward_on_call", "palliative_on_call", "radiation_on_call"}),
    ("מחקר", {"research"}),
    ("יום עיון / כנס", {"study_day"}),
]


def summarize_schedule(records: pd.DataFrame) -> pd.DataFrame:
    if records.empty:
        return pd.DataFrame()
    employees = sorted(records["employee"].dropna().unique())
    rows: list[dict[str, Any]] = []
    duty_codes = {"ward_duty_regular", "ward_duty_friday", "ward_duty_saturday", "er_duty", "day_hospital_duty"}
    for employee in employees:
        employee_records = records[records["employee"] == employee]
        tasks = employee_records[employee_records["record_type"] == "task"]
        statuses = employee_records[employee_records["record_type"] == "status"]
        row: dict[str, Any] = {"עובד/ת": employee}
        for label, codes in SUMMARY_DEFINITIONS:
            row[label] = int(tasks["task_code"].isin(codes).sum())
        row["אחרי תורנות"] = int((statuses["task_code"] == "after_duty").sum())
        row["חופש"] = int((statuses["task_code"] == "vacation").sum())
        row["היעדרות / אחר"] = int(statuses["task_code"].isin({"absence", "absence_other"}).sum())
        row["סה\"כ תורנויות"] = int(tasks["task_code"].isin(duty_codes).sum())
        holiday_mask = tasks["holiday"].astype(str).str.strip() != ""
        row["תורנויות בחג / יום מיוחד"] = int((tasks["task_code"].isin(duty_codes) & holiday_mask).sum())
        row["מטלות בחג / יום מיוחד"] = int(holiday_mask.sum())
        weekend_or_holiday = tasks["date"].apply(lambda value: value.weekday() in {4, 5}) | holiday_mask
        row["שיבוצים בסופ\"ש / חג"] = int(weekend_or_holiday.sum())
        row["סה\"כ מטלות"] = int(len(tasks))
        rows.append(row)
    return pd.DataFrame(rows)


def validate_schedule(records: pd.DataFrame, config: Mapping[str, Any]) -> pd.DataFrame:
    """Return only coverage problems found in the parsed schedule.

    The numeric requirements are read from configuration so the engine remains
    generic and the department can change them without editing parsing logic.
    """
    columns = ["תאריך", "יום", "חג / יום מיוחד", "בדיקה", "נדרש", "נמצא", "פירוט"]
    if records.empty:
        return pd.DataFrame(columns=columns)

    rules = config.get("staffing_rules", {})
    expected_department = {
        "weekday": int(rules.get("weekday_department", 3)),
        "friday": int(rules.get("friday_department", 2)),
        "saturday": int(rules.get("saturday_department", 2)),
    }
    expected_ward_duty = int(rules.get("ward_duty_per_day", 1))
    tasks = records[records["record_type"] == "task"].copy()
    issues: list[dict[str, Any]] = []

    department_codes = {
        "department_regular",
        "department_weekend",
        "department_weekend_friday_duty",
        "friday_visit",
        "department_weekend_outgoing",
        "department_weekend_saturday_duty",
    }
    ward_duty_codes = {"ward_duty_regular", "ward_duty_friday", "ward_duty_saturday"}

    configured_dates = records.attrs.get("schedule_dates", [])
    dates_to_check = sorted(set(configured_dates) or set(tasks["date"].dropna().tolist()))
    for current in dates_to_check:
        current_date = current.date() if isinstance(current, pd.Timestamp) else current
        day_tasks = tasks[tasks["date"] == current_date]
        holiday_values = [normalize_spaces(value) for value in day_tasks["holiday"].tolist() if normalize_spaces(value)]
        holiday = holiday_values[0] if holiday_values else important_day_name(current_date)
        if current_date.weekday() == 4:
            day_type = "friday"
        elif current_date.weekday() == 5:
            day_type = "saturday"
        else:
            day_type = "weekday"

        department_tasks = day_tasks[day_tasks["task_code"].isin(department_codes)]
        department_names = sorted(set(department_tasks["employee"].dropna()))
        expected = expected_department[day_type]
        if len(department_names) != expected:
            issues.append(
                {
                    "תאריך": current_date,
                    "יום": hebrew_weekday(current_date),
                    "חג / יום מיוחד": holiday,
                    "בדיקה": "איוש מחלקה",
                    "נדרש": expected,
                    "נמצא": len(department_names),
                    "פירוט": ", ".join(department_names) if department_names else "לא נמצא שיבוץ",
                }
            )

        duty_tasks = day_tasks[day_tasks["task_code"].isin(ward_duty_codes)]
        duty_names = sorted(set(duty_tasks["employee"].dropna()))
        if len(duty_names) != expected_ward_duty:
            issues.append(
                {
                    "תאריך": current_date,
                    "יום": hebrew_weekday(current_date),
                    "חג / יום מיוחד": holiday,
                    "בדיקה": "תורן מחלקה",
                    "נדרש": expected_ward_duty,
                    "נמצא": len(duty_names),
                    "פירוט": ", ".join(duty_names) if duty_names else "לא נמצא תורן",
                }
            )

        if day_type == "friday":
            visit_names = sorted(set(day_tasks.loc[day_tasks["task_code"] == "friday_visit", "employee"].dropna()))
            if len(visit_names) != 1:
                issues.append(
                    {
                        "תאריך": current_date,
                        "יום": hebrew_weekday(current_date),
                        "חג / יום מיוחד": holiday,
                        "בדיקה": "ביקור שישי נוסף",
                        "נדרש": 1,
                        "נמצא": len(visit_names),
                        "פירוט": ", ".join(visit_names) if visit_names else "לא נמצא עובד נוסף",
                    }
                )

        if day_type == "saturday":
            outgoing_names = sorted(set(day_tasks.loc[day_tasks["task_code"] == "department_weekend_outgoing", "employee"].dropna()))
            saturday_names = sorted(set(day_tasks.loc[day_tasks["task_code"] == "department_weekend_saturday_duty", "employee"].dropna()))
            for check_name, names in [("תורן יוצא של שישי", outgoing_names), ("תורן שבת במחלקה", saturday_names)]:
                if len(names) != 1:
                    issues.append(
                        {
                            "תאריך": current_date,
                            "יום": hebrew_weekday(current_date),
                            "חג / יום מיוחד": holiday,
                            "בדיקה": check_name,
                            "נדרש": 1,
                            "נמצא": len(names),
                            "פירוט": ", ".join(names) if names else "לא נמצא שיבוץ",
                        }
                    )

    return pd.DataFrame(issues, columns=columns)
