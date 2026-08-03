from __future__ import annotations

from io import BytesIO, StringIO
from typing import Any, Iterable, Mapping, Sequence

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from core import (
    HEBREW_MONTHS,
    availability_matrix,
    availability_summary,
    build_month_table,
    decode_submission,
    normalize_spaces,
    submissions_to_long_table,
)

DARK_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
BASE_FILL = PatternFill("solid", fgColor="E2F0D9")
WEEKEND_FILL = PatternFill("solid", fgColor="FFF2CC")
HOLIDAY_FILL = PatternFill("solid", fgColor="FCE4D6")
BLOCKED_FILL = PatternFill("solid", fgColor="D9D9D9")
VACATION_FILL = PatternFill("solid", fgColor="F4CCCC")
NOTE_FILL = PatternFill("solid", fgColor="DDEBF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_GRAY = Side(style="thin", color="D9E1F2")


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _setup_sheet(sheet, freeze_panes: str | None = None) -> None:
    sheet.sheet_view.rightToLeft = True
    sheet.sheet_view.showGridLines = False
    if freeze_panes:
        sheet.freeze_panes = freeze_panes


def _apply_header(cell, dark: bool = True) -> None:
    cell.fill = DARK_FILL if dark else GROUP_FILL
    cell.font = WHITE_FONT if dark else BOLD_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_dataframe(sheet, dataframe: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    for col_offset, column in enumerate(dataframe.columns, start=start_col):
        cell = sheet.cell(start_row, col_offset, str(column))
        _apply_header(cell)

    for row_offset, row in enumerate(dataframe.itertuples(index=False), start=start_row + 1):
        for col_offset, value in enumerate(row, start=start_col):
            cell = sheet.cell(row_offset, col_offset)
            if pd.isna(value):
                value = ""
            if hasattr(value, "to_pydatetime"):
                value = value.to_pydatetime()
            cell.value = value
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
                cell.number_format = "dd.mm.yyyy"


def build_personal_submission_workbook(
    employee: str,
    edited_table: pd.DataFrame,
    encoded_submission: str,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "העדפות אישיות"
    _setup_sheet(sheet, "A2")

    export = edited_table.copy()
    export["תאריך"] = pd.to_datetime(export["תאריך"]).dt.date
    _write_dataframe(sheet, export)
    sheet.auto_filter.ref = sheet.dimensions

    widths = {"A": 13, "B": 8, "C": 24, "D": 12, "E": 12, "F": 36}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    payload_sheet = workbook.create_sheet("פלט להעתקה")
    _setup_sheet(payload_sheet)
    payload_sheet["A1"] = "שם עובד/ת"
    payload_sheet["B1"] = employee
    payload_sheet["A3"] = "יש להעתיק את כל הטקסט שבתא B3 לקובץ הריכוז"
    payload_sheet["B3"] = encoded_submission
    payload_sheet["B3"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    payload_sheet.column_dimensions["A"].width = 34
    payload_sheet.column_dimensions["B"].width = 100
    payload_sheet.row_dimensions[3].height = 120
    return workbook_bytes(workbook)


def build_response_collection_template(year: int, month: int, rows: int = 50) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ריכוז תשובות"
    _setup_sheet(sheet, "A2")
    sheet["A1"] = "שם עובד/ת"
    sheet["B1"] = "תשובה להדבקה"
    sheet["C1"] = "הערת מתכנן"
    for cell in sheet[1]:
        _apply_header(cell)
    for row in range(2, rows + 2):
        sheet.cell(row, 1).alignment = Alignment(horizontal="right")
        sheet.cell(row, 2).alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        sheet.cell(row, 3).alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        sheet.row_dimensions[row].height = 45
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 100
    sheet.column_dimensions["C"].width = 36
    sheet.auto_filter.ref = f"A1:C{rows + 1}"

    instructions = workbook.create_sheet("הנחיות")
    _setup_sheet(instructions)
    instructions["A1"] = f"ריכוז העדפות לחודש {HEBREW_MONTHS[month]} {year}"
    instructions["A1"].font = Font(bold=True, size=16)
    instructions["A3"] = "1. בכל שורה מדביקים פלט מלא של עובד/ת אחד/ת בעמודה 'תשובה להדבקה'."
    instructions["A4"] = "2. ניתן למלא את השם גם בעמודה הראשונה לצורך בקרה."
    instructions["A5"] = "3. מעלים את הקובץ לכלי 2 באפליקציה."
    instructions.column_dimensions["A"].width = 110
    return workbook_bytes(workbook)


def parse_response_collection(file_bytes: bytes, filename: str) -> tuple[list[dict[str, Any]], list[str]]:
    filename_lower = filename.lower()
    warnings: list[str] = []
    payloads: list[dict[str, Any]] = []

    if filename_lower.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig")
        dataframe = pd.read_csv(StringIO(text))
    else:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = workbook["ריכוז תשובות"] if "ריכוז תשובות" in workbook.sheetnames else workbook.active
        header_row = None
        response_col = None
        name_col = None
        for row in range(1, min(sheet.max_row, 12) + 1):
            values = [normalize_spaces(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
            for index, value in enumerate(values, start=1):
                if value in {"תשובה להדבקה", "תשובה", "פלט", "פלט להעתקה"}:
                    header_row, response_col = row, index
                if value in {"שם עובד/ת", "שם", "עובד/ת", "עובד"}:
                    name_col = index
            if response_col:
                break
        if not response_col or not header_row:
            raise ValueError("לא נמצאה עמודה בשם 'תשובה להדבקה'")

        rows = []
        for row in range(header_row + 1, sheet.max_row + 1):
            response = sheet.cell(row, response_col).value
            if not normalize_spaces(response):
                continue
            rows.append(
                {
                    "תשובה להדבקה": response,
                    "שם עובד/ת": sheet.cell(row, name_col).value if name_col else "",
                    "שורת מקור": row,
                }
            )
        dataframe = pd.DataFrame(rows)

    normalized_columns = {normalize_spaces(column): column for column in dataframe.columns}
    response_source = next(
        (normalized_columns[key] for key in ["תשובה להדבקה", "תשובה", "פלט", "פלט להעתקה"] if key in normalized_columns),
        None,
    )
    name_source = next(
        (normalized_columns[key] for key in ["שם עובד/ת", "שם", "עובד/ת", "עובד"] if key in normalized_columns),
        None,
    )
    if response_source is None:
        raise ValueError("לא נמצאה עמודה בשם 'תשובה להדבקה'")

    for index, row in dataframe.iterrows():
        raw = row.get(response_source)
        if not normalize_spaces(raw):
            continue
        try:
            payload = decode_submission(raw)
            typed_name = normalize_spaces(row.get(name_source)) if name_source else ""
            payload_name = normalize_spaces(payload.get("employee"))
            if typed_name and typed_name != payload_name:
                warnings.append(f"שורה {index + 2}: השם בעמודה ({typed_name}) שונה מהשם בפלט ({payload_name}); נעשה שימוש בשם שבפלט.")
            payloads.append(payload)
        except Exception as exc:
            warnings.append(f"שורה {index + 2}: {exc}")

    deduplicated: dict[str, dict[str, Any]] = {}
    for payload in payloads:
        employee = normalize_spaces(payload.get("employee"))
        if employee in deduplicated:
            warnings.append(f"נמצאה יותר מתשובה אחת עבור {employee}; נעשה שימוש בתשובה האחרונה.")
        deduplicated[employee] = payload
    return list(deduplicated.values()), warnings


def build_schedule_template(year: int, month: int, config: Mapping[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "סידור עבודה"
    _setup_sheet(sheet, "A4")

    final_column = 25  # Y
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=final_column)
    title = sheet.cell(1, 1, f"{HEBREW_MONTHS[month]} {year}")
    title.fill = DARK_FILL
    title.font = Font(color="FFFFFF", bold=True, size=18)
    title.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 8

    headers = [
        (1, 1, "תאריך"),
        (2, 2, "יום"),
        (3, 3, "חג / יום מיוחד"),
        (4, 6, "מעקב אישי"),
        (7, 9, "מחלקה"),
        (10, 10, "ATT"),
        (11, 11, "א.יום"),
        (12, 12, "מיון"),
        (13, 13, "אבחון מהיר"),
        (14, 14, "תגבור מיון"),
        (15, 18, "מרפאות"),
        (19, 19, "כונן מחלקה"),
        (20, 20, "כונן פליאציה"),
        (21, 21, "כונן קרינה"),
        (22, 22, "ת. א.יום (21:00)"),
        (23, 23, "תורן מיון"),
        (24, 24, "תורן מחלקה"),
        (25, 25, "היעדרויות / אחר"),
    ]
    for start_col, end_col, label in headers:
        for column in range(start_col, end_col + 1):
            header_cell = sheet.cell(3, column)
            _apply_header(header_cell, dark=start_col <= 3 or start_col >= 22)
        sheet.cell(3, start_col).value = label
        if start_col != end_col:
            sheet.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)

    month_table = build_month_table(year, month)
    for row_index, row in enumerate(month_table.itertuples(index=False), start=4):
        current = row[0]
        weekday = row[1]
        holiday = row[2]
        sheet.cell(row_index, 1, current)
        sheet.cell(row_index, 1).number_format = "dd.mm.yyyy"
        sheet.cell(row_index, 2, weekday)
        sheet.cell(row_index, 3, holiday)
        for column in range(1, final_column + 1):
            cell = sheet.cell(row_index, column)
            cell.alignment = Alignment(horizontal="center" if column <= 3 else "right", vertical="center", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)
            if current.weekday() in {4, 5}:
                cell.fill = WEEKEND_FILL
            if holiday:
                cell.fill = HOLIDAY_FILL
        sheet.row_dimensions[row_index].height = 42

    widths = {
        1: 13, 2: 7, 3: 22,
        4: 15, 5: 15, 6: 15,
        7: 15, 8: 15, 9: 15,
        10: 11, 11: 14, 12: 14, 13: 16, 14: 14,
        15: 17, 16: 17, 17: 17, 18: 17,
        19: 16, 20: 16, 21: 16,
        22: 17, 23: 16, 24: 16, 25: 26,
    }
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.auto_filter.ref = f"A3:Y{sheet.max_row}"
    return workbook_bytes(workbook)


def _style_availability_sheet(sheet, dataframe: pd.DataFrame) -> None:
    _setup_sheet(sheet, "D2")
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        _apply_header(cell)
    for row in range(2, sheet.max_row + 1):
        current_date = sheet.cell(row, 1).value
        holiday = normalize_spaces(sheet.cell(row, 3).value)
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            if hasattr(current_date, "weekday") and current_date.weekday() in {4, 5}:
                cell.fill = WEEKEND_FILL
            if holiday:
                cell.fill = HOLIDAY_FILL
            if col >= 4:
                text = normalize_spaces(cell.value)
                if "חופש" in text:
                    cell.fill = VACATION_FILL
                elif "לא זמין" in text:
                    cell.fill = BLOCKED_FILL
                elif text:
                    cell.fill = NOTE_FILL
    sheet.column_dimensions["A"].width = 13
    sheet.column_dimensions["B"].width = 7
    sheet.column_dimensions["C"].width = 24
    for col in range(4, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 19


def append_availability_sheets(
    base_workbook_bytes: bytes,
    payloads: Sequence[Mapping[str, Any]],
    year: int,
    month: int,
) -> bytes:
    workbook = load_workbook(BytesIO(base_workbook_bytes))
    for name in ["זמינות צוות", "נתוני העדפות", "סיכום העדפות"]:
        if name in workbook.sheetnames:
            del workbook[name]

    matrix = availability_matrix(payloads, year, month)
    normalized = submissions_to_long_table(payloads)
    summary = availability_summary(payloads)

    matrix_sheet = workbook.create_sheet("זמינות צוות")
    _write_dataframe(matrix_sheet, matrix)
    _style_availability_sheet(matrix_sheet, matrix)

    data_sheet = workbook.create_sheet("נתוני העדפות")
    _setup_sheet(data_sheet, "A2")
    _write_dataframe(data_sheet, normalized)
    data_sheet.auto_filter.ref = data_sheet.dimensions

    summary_sheet = workbook.create_sheet("סיכום העדפות")
    _setup_sheet(summary_sheet, "A2")
    _write_dataframe(summary_sheet, summary)
    summary_sheet.auto_filter.ref = summary_sheet.dimensions

    return workbook_bytes(workbook)


def build_manager_workbook(
    payloads: Sequence[Mapping[str, Any]],
    year: int,
    month: int,
    config: Mapping[str, Any],
    uploaded_template: bytes | None = None,
) -> bytes:
    base = uploaded_template or build_schedule_template(year, month, config)
    return append_availability_sheets(base, payloads, year, month)


def build_schedule_analysis_workbook(
    records: pd.DataFrame,
    summary: pd.DataFrame,
    events: pd.DataFrame,
    validation: pd.DataFrame | None = None,
) -> bytes:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    sheets = [("מטלות מפוענחות", records), ("סיכום מטלות", summary), ("בדיקות איוש", validation if validation is not None else pd.DataFrame()), ("אירועי יומן", events)]
    for name, dataframe in sheets:
        sheet = workbook.create_sheet(name)
        _setup_sheet(sheet, "A2")
        _write_dataframe(sheet, dataframe)
        sheet.auto_filter.ref = sheet.dimensions
        for column in range(1, sheet.max_column + 1):
            max_length = max(
                [len(normalize_spaces(sheet.cell(row, column).value)) for row in range(1, min(sheet.max_row, 200) + 1)] + [10]
            )
            sheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 12), 45)
    return workbook_bytes(workbook)
